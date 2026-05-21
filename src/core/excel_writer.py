import os

import openpyxl

from src.core.models import MaterialGroup
from src.core.profiles import MaterialProfileParser


class ExcelWorkbookWriter:
    def __init__(self, caminho_planilha: str, parser: MaterialProfileParser | None = None):
        self.caminho_planilha = caminho_planilha
        self.parser = parser or MaterialProfileParser()
        self.workbook = openpyxl.load_workbook(caminho_planilha)
        self.sheet = self.workbook.active
        nome_base, extensao = os.path.splitext(caminho_planilha)
        self.caminho_processado = f"{nome_base}_processado{extensao}"

    def encontrar_proxima_linha_vazia(self, codigo_secao: str, linha_inicio_busca: int) -> int | None:
        for row in range(linha_inicio_busca, self.sheet.max_row + 2):
            celula_codigo = self.sheet.cell(row=row, column=1)
            if str(celula_codigo.value).strip() != codigo_secao:
                continue

            if self.sheet.cell(row=row, column=10).value in [None, 0, "", "0", 0.0]:
                return row

        return None

    def encontrar_linha_viga_w(self, nome_normalizado: str) -> int | None:
        alvo = nome_normalizado.replace(" ", "")
        for row in range(4, self.sheet.max_row + 1):
            val_cell = str(self.sheet.cell(row=row, column=1).value).strip().upper().replace(" ", "")
            if val_cell == alvo and self.sheet.cell(row=row, column=10).value in [None, 0, "", 0.0]:
                return row
        return None

    def escrever_grupo(self, linha_alvo: int, grupo: MaterialGroup) -> None:
        dim_a, dim_b, dim_c, dim_esp = self.parser.parse_dimensoes_inteligente(grupo.descricao, grupo.tipo_perfil)

        if grupo.tipo_perfil != "VIGA_W":
            if grupo.tipo_perfil in ["PERFIL_U", "TERCA"]:
                self.sheet.cell(row=linha_alvo, column=2).value = dim_a
                self.sheet.cell(row=linha_alvo, column=4).value = dim_b
                self.sheet.cell(row=linha_alvo, column=6).value = dim_c
            elif grupo.tipo_perfil == "CANTONEIRA":
                self.sheet.cell(row=linha_alvo, column=4).value = dim_a
                self.sheet.cell(row=linha_alvo, column=6).value = dim_b

            self.sheet.cell(row=linha_alvo, column=8).value = dim_esp

        self.sheet.cell(row=linha_alvo, column=9).value = grupo.aco
        if grupo.comprimento_m > 0:
            self.sheet.cell(row=linha_alvo, column=10).value = grupo.comprimento_m
        self.sheet.cell(row=linha_alvo, column=17).value = grupo.peso_total

    def ocultar_linhas_vazias(self, linha_inicio: int = 4) -> None:
        for row in range(linha_inicio, self.sheet.max_row + 1):
            valor_a = str(self.sheet.cell(row=row, column=1).value).strip().upper()

            if "TOTAL" in valor_a or "ATIVO FINAL" in valor_a or "RESUMO" in valor_a:
                for row_resumo in range(row, min(row + 25, self.sheet.max_row + 1)):
                    self.sheet.row_dimensions[row_resumo].hidden = False
                break

            valor_comprimento = self.sheet.cell(row=row, column=10).value
            if self.sheet.cell(row=row, column=1).value and valor_comprimento in [None, 0, "0", "", 0.0]:
                self.sheet.row_dimensions[row].hidden = True
            else:
                self.sheet.row_dimensions[row].hidden = False

    def salvar(self) -> str:
        self.workbook.save(self.caminho_processado)
        return self.caminho_processado
